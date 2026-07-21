from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
import io
import os
from typing import Generator
import uuid

from azure.core.exceptions import HttpResponseError, ResourceExistsError
from azure.storage.blob import BlobSasPermissions, BlobServiceClient, generate_blob_sas
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
import psycopg2.extras
import psycopg2.pool

load_dotenv()

# CONFIG (pulled from environment / Azure App Settings)
AZURE_STORAGE_CONNECTION_STRING = os.environ["AZURE_STORAGE_CONNECTION_STRING"]
AZURE_STORAGE_ACCOUNT_NAME = os.environ["AZURE_STORAGE_ACCOUNT_NAME"]
AZURE_STORAGE_ACCOUNT_KEY = os.environ["AZURE_STORAGE_ACCOUNT_KEY"]

AZURE_CONTAINER_FRI = os.environ.get("AZURE_CONTAINER_FRI", "fri-exports")
AZURE_CONTAINER_MNRL = os.environ.get("AZURE_CONTAINER_MNRL", "mnrl-exports")
AZURE_CONTAINER_MATCHOFF_OUTPUT = os.environ.get("AZURE_CONTAINER_MATCHOFF_OUTPUT", "matchoff-output")

PG_DSN = os.environ["AZURE_POSTGRES_DSN"]
PG_MATCHOFF_TABLE = os.environ.get("PG_MATCHOFF_TABLE", "mobilecleanup")
PG_MOB_COLUMN = os.environ.get("PG_MOB_COLUMN", "final_mob_num")
PG_CUSTID_COLUMN = os.environ.get("PG_CUSTID_COLUMN", "cust_id")

# Matchoff tuning
SQL_CHUNK_SIZE = int(os.environ.get("MATCHOFF_SQL_CHUNK", "1000"))
SQL_MAX_WORKERS = int(os.environ.get("MATCHOFF_SQL_WORKERS", "5"))
EXCEL_FLUSH_ROWS = int(os.environ.get("MATCHOFF_EXCEL_FLUSH", "5000"))

FRI_STATUS_COLUMN = "Status"
FRI_EXCLUDE_STATUS = "compliant"

# Excel row limit Excel hard cap is 1,048,576 rows (incl. header).
EXCEL_MAX_ROWS = int(os.environ.get("EXCEL_MAX_ROWS", "1_000_000"))

# Blob chunked-upload threshold - files larger than this use streams
BLOB_CHUNK_THRESHOLD = 64 * 1024 * 1024  # 64 MB

# Valid source containers
VALID_SOURCE_CONTAINERS = {
    "fri-exports": AZURE_CONTAINER_FRI,
    "mnrl-exports": AZURE_CONTAINER_MNRL,
}

MOBILECLEANUP_FETCH_COLUMNS: list[str] = [
    "cust_id",
    "final_mob_num",
    "zbadid",
    "zbnad1",
    "zbnad2",
    "zbnad3",
    "zbnad4",
    "zbnad5",
    "zbteno",
    "zbten2",
    "zbpscd",
    "zbiadd",
]

MOBILECLEANUP_Headers: list[str] = [
    "Customer Id",
    "Mobile Phone",
    "Address ID",
    "Customer name & address",
    "Customer name & address",
    "Customer name & address",
    "Customer name & address",
    "Customer name & address",
    "Telephone No",
    "Telephone No 2",
    "Postal code",
    "Internet Address",
]

_pg_pool: psycopg2.pool.ThreadedConnectionPool | None = None


def _get_pool() -> psycopg2.pool.ThreadedConnectionPool:
    global _pg_pool
    if _pg_pool is None or _pg_pool.closed:
        _pg_pool = psycopg2.pool.ThreadedConnectionPool(
            minconn=2,
            maxconn=max(SQL_MAX_WORKERS + 2, 10),
            dsn=PG_DSN,
        )
    return _pg_pool


def _pg_conn():
    """Obtain a connection from the pool.

    IMPORTANT: callers must return it via _put_conn() when done. Use the
    context manager _pg_cursor() for query blocks.
    """
    return _get_pool().getconn()


def _put_conn(conn):
    try:
        _get_pool().putconn(conn)
    except Exception:
        pass


class _pg_cursor:
    """Context manager that borrows a connection from the pool, yields a

    RealDictCursor, and returns the connection on exit.
    """

    def __init__(self, dict_cursor: bool = False):
        self.dict_cursor = dict_cursor
        self._conn = None
        self._cur = None

    def __enter__(self):
        self._conn = _pg_conn()
        factory = psycopg2.extras.RealDictCursor if self.dict_cursor else None
        self._cur = self._conn.cursor(cursor_factory=factory) if factory else self._conn.cursor()
        return self._cur

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            try:
                self._conn.rollback()
            except Exception:
                pass
        else:
            try:
                self._conn.commit()
            except Exception:
                pass
        try:
            if self._cur:
                self._cur.close()
        except Exception:
            pass
        if self._conn:
            _put_conn(self._conn)
        return False  # re-raise exceptions


# INTERNAL UTILITIES
def _ensure_container(name: str) -> None:
    try:
        cc = _blob_service().get_container_client(name)
        cc.create_container()
        print(f"[BLOB] Created Container '{name}'")
    except ResourceExistsError:
        pass
    except HttpResponseError as e:
        raise RuntimeError(f"Failed to ensure container '{name}': {e.message}") from e


def _ensure_all_containers() -> None:
    for name in (AZURE_CONTAINER_FRI, AZURE_CONTAINER_MNRL, AZURE_CONTAINER_MATCHOFF_OUTPUT):
        _ensure_container(name)


def _get_container_name(module: str) -> str:
    return AZURE_CONTAINER_FRI if module.upper() == "FRI" else AZURE_CONTAINER_MNRL


def _blob_service() -> BlobServiceClient:
    return BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)


def _chunks(lst: list, n: int) -> Generator:
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i : i + n]


# EXCEL BUILDERS
def _make_header_cells(ws, columns: list) -> list:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_align = Alignment(horizontal="center")
    cells = []
    for col in columns:
        cell = WriteOnlyCell(ws, value=col.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cells.append(cell)
    return cells


def build_excel(records: list, columns: list) -> io.BytesIO:
    buf = io.BytesIO()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Data")
    ws.append(_make_header_cells(ws, columns))

    for chunk in _chunks(records, 5_000):
        for rec in chunk:
            ws.append([rec.get(c, "") for c in columns])

    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_parts(
    records: list,
    base_file_name: str,
    columns: list,
    max_rows: int = EXCEL_MAX_ROWS,
) -> list[tuple[io.BytesIO, str, int]]:
    total_parts = -(-len(records) // max_rows)  # ceiling division
    results: list[tuple[io.BytesIO, str, int]] = []

    for part_idx, chunk in enumerate(_chunks(records, max_rows), start=1):
        if total_parts == 1:
            fname = base_file_name
        else:
            stem = base_file_name.replace(".xlsx", "")
            fname = f"{stem}_part{part_idx}.xlsx"

        buf = io.BytesIO()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Data")
        ws.append(_make_header_cells(ws, columns))

        for row_chunk in _chunks(chunk, 5_000):
            for rec in row_chunk:
                ws.append([rec.get(c, "") for c in columns])

        wb.save(buf)
        buf.seek(0)
        results.append((buf, fname, len(chunk)))

    return results


def build_matched_excel(
    input_rows: list[dict],
    headers: list[str],
    mob_col: str,
    matched_map: dict[str, dict],
) -> tuple[io.BytesIO, int]:
    output_headers = headers + MOBILECLEANUP_Headers
    buf = io.BytesIO()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Matched Data")
    ws.append(_make_header_cells(ws, output_headers))

    matched_count = 0
    buffer: list = []

    for row in input_rows:
        mob_val = str(row.get(mob_col, "")).strip()
        if mob_val not in matched_map:
            continue

        data_row = [row.get(h, "") for h in headers]
        col_data = matched_map[mob_val]
        data_row += [col_data.get(col, "") for col in MOBILECLEANUP_FETCH_COLUMNS]

        buffer.append(data_row)
        matched_count += 1

        if len(buffer) >= EXCEL_FLUSH_ROWS:
            for r in buffer:
                ws.append(r)
            buffer.clear()

    for r in buffer:
        ws.append(r)

    wb.save(buf)
    buf.seek(0)
    return buf, matched_count


# AZURE BLOB CORE OPERATIONS
def upload_blob(container: str, blob_name: str, data: bytes | io.BytesIO) -> str:
    _ensure_container(container)
    blob_client = _blob_service().get_blob_client(container=container, blob=blob_name)

    if isinstance(data, (bytes, bytearray)):
        size = len(data)
        stream = io.BytesIO(data)
    else:
        data.seek(0, 2)
        size = data.tell()
        data.seek(0)
        stream = data

    if size > BLOB_CHUNK_THRESHOLD:
        blob_client.upload_blob(stream, overwrite=True, max_concurrency=4)
    else:
        blob_client.upload_blob(stream, overwrite=True)

    return blob_client.url


def get_sas_download_url(container: str, blob_name: str, expiry_hours: int = 1) -> str:
    sas_token = generate_blob_sas(
        account_name=AZURE_STORAGE_ACCOUNT_NAME,
        container_name=container,
        blob_name=blob_name,
        account_key=AZURE_STORAGE_ACCOUNT_KEY,
        permission=BlobSasPermissions(read=True),
        expiry=datetime.now(timezone.utc) + timedelta(hours=expiry_hours),
    )
    return f"https://{AZURE_STORAGE_ACCOUNT_NAME}.blob.core.windows.net/{container}/{blob_name}?{sas_token}"


# BLOB BROWSER HELPERS
def list_source_date_folders(container: str) -> list[str]:
    if container not in VALID_SOURCE_CONTAINERS.values():
        raise ValueError(f"Container '{container}' is not a permitted source.")

    cc = _blob_service().get_container_client(container)
    seen = set()
    for blob in cc.list_blobs():
        parts = blob.name.split("/")
        if len(parts) >= 2 and parts[0]:
            seen.add(parts[0])

    return sorted(seen, reverse=True)


def list_source_files(container: str, date_folder: str) -> list[str]:
    if container not in VALID_SOURCE_CONTAINERS.values():
        raise ValueError(f"Container '{container}' is not a permitted source.")

    prefix = f"{date_folder}/"
    cc = _blob_service().get_container_client(container)
    files = []
    for blob in cc.list_blobs(name_starts_with=prefix):
        fname = blob.name[len(prefix) :]
        if fname and fname.endswith(".xlsx") and "/" not in fname:
            files.append(fname)

    return sorted(files)


# BLOB READ source Excel files
def read_excel_from_blob(container: str, date_folder: str, file_name: str) -> tuple[list[dict], list[str]]:
    if container not in VALID_SOURCE_CONTAINERS.values():
        raise ValueError(f"Container '{container}' is not a permitted source.")

    blob_path = f"{date_folder}/{file_name}"
    bc = _blob_service().get_blob_client(container=container, blob=blob_path)

    try:
        stream = io.BytesIO()
        bc.download_blob().readinto(stream)
        stream.seek(0)
    except Exception as e:
        raise ValueError(f"Could not read blob '{blob_path}' from '{container}': {e}") from e

    wb = load_workbook(filename=stream, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError(f"Excel file '{file_name}' is empty.")

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
    rows = [dict(zip(headers, row)) for row in rows_iter]
    wb.close()

    if not rows:
        raise ValueError(f"Excel file '{file_name}' has a header row but no data rows.")

    return rows, headers


# BLOB UPLOAD matchoff output
def upload_matchoff_output(buf: io.BytesIO, date_folder: str, file_name: str) -> str:
    _ensure_container(AZURE_CONTAINER_MATCHOFF_OUTPUT)
    blob_path = f"{date_folder}/{file_name}"
    bc = _blob_service().get_blob_client(container=AZURE_CONTAINER_MATCHOFF_OUTPUT, blob=blob_path)
    bc.upload_blob(buf, overwrite=True)
    return blob_path


# MOBILECLEANUP chunked parallel matching
def _fetch_chunk(mob_chunk: list[str]) -> dict[str, dict]:
    if not mob_chunk:
        return {}

    all_cols = list(dict.fromkeys([PG_MOB_COLUMN] + MOBILECLEANUP_FETCH_COLUMNS))
    select_cols = ", ".join(all_cols)
    sql = f"SELECT {select_cols} FROM {PG_MATCHOFF_TABLE} WHERE {PG_MOB_COLUMN} IN %s"

    result = {}
    conn = _pg_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, (tuple(mob_chunk),))
            col_names = [desc[0] for desc in cur.description]
            for row in cur.fetchall():
                mob_val = str(row[0]).strip()
                col_data = {col_names[i]: (str(val).strip() if val is not None else "") for i, val in enumerate(row)}
                result[mob_val] = col_data
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        _put_conn(conn)

    return result


def match_mob_numbers(
    mob_numbers: list,
    chunk_size: int = SQL_CHUNK_SIZE,
    max_workers: int = SQL_MAX_WORKERS,
) -> dict[str, dict]:
    unique_mobs = list({str(m).strip() for m in mob_numbers if m is not None and str(m).strip()})
    if not unique_mobs:
        return {}

    chunks = list(_chunks(unique_mobs, chunk_size))
    matched: dict[str, dict] = {}

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(_fetch_chunk, chunk): chunk for chunk in chunks}
        for future in as_completed(futures):
            try:
                matched.update(future.result())
            except Exception as exc:
                print(f"[MATCHOFF] Chunk query failed: {exc}")

    return matched


# POSTGRESQL export_files CRUD
def save_export_metadata(
    module: str,
    file_name: str,
    blob_name: str,
    record_count: int,
    export_date: str,
    data_type: str = None,
) -> str:
    """Insert export metadata row. Returns the new UUID string."""
    row_id = str(uuid.uuid4())
    container = _get_container_name(module)
    with _pg_cursor() as cur:
        cur.execute(
            """
            INSERT INTO export_files
            (id, module, file_name, blob_name, container, record_count, export_date, data_type, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """,
            (row_id, module.upper(), file_name, blob_name, container, record_count, export_date, data_type),
        )
    return row_id


def list_export_metadata(module: str) -> list:
    with _pg_cursor(dict_cursor=True) as cur:
        cur.execute(
            """
            SELECT id, file_name, blob_name, container, record_count, export_date, data_type, created_at
            FROM export_files
            WHERE module = %s
            ORDER BY created_at DESC
            """,
            (module.upper(),),
        )
        rows = cur.fetchall()

    result = []
    for row in rows:
        d = dict(row)
        d["created_at"] = d["created_at"].isoformat()
        d["export_date"] = str(d["export_date"])
        result.append(d)

    return result


def get_export_metadata_row(file_id: str, module: str) -> dict:
    """Single export row by UUID + module. Raises ValueError if not found."""
    with _pg_cursor(dict_cursor=True) as cur:
        cur.execute(
            "SELECT * FROM export_files WHERE id = %s AND module = %s",
            (file_id, module.upper()),
        )
        row = cur.fetchone()

    if not row:
        raise ValueError(f"No export record found: id={file_id}, module={module}")
    return dict(row)


# POSTGRESQL - matchoff_runs CRUD
def save_matchoff_metadata(
    source_container: str,
    source_date_folder: str,
    source_file_name: str,
    output_blob_path: str,
    input_row_count: int,
    matched_count: int,
    mob_col: str,
    run_date: str,
) -> str:
    row_id = str(uuid.uuid4())
    with _pg_cursor() as cur:
        cur.execute(
            """
            INSERT INTO matchoff_runs (
                id, source_container, source_date_folder, source_file_name,
                output_blob_path, output_container, input_row_count,
                matched_count, mob_column, run_date, created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """,
            (
                row_id,
                source_container,
                source_date_folder,
                source_file_name,
                output_blob_path,
                AZURE_CONTAINER_MATCHOFF_OUTPUT,
                input_row_count,
                matched_count,
                mob_col,
                run_date,
            ),
        )
    return row_id


def list_matchoff_metadata() -> list:
    """All matchoff runs, newest first."""
    with _pg_cursor(dict_cursor=True) as cur:
        cur.execute(
            """
            SELECT id, source_container, source_date_folder, source_file_name,
                   output_blob_path, output_container, input_row_count,
                   matched_count, mob_column, run_date, created_at
            FROM matchoff_runs
            ORDER BY created_at DESC
            """
        )
        rows = cur.fetchall()

    result = []
    for row in rows:
        d = dict(row)
        d["created_at"] = d["created_at"].isoformat()
        d["run_date"] = str(d["run_date"])
        result.append(d)

    return result


def get_matchoff_run(run_id: str) -> dict:
    """Single matchoff row by UUID. Raises ValueError if not found."""
    with _pg_cursor(dict_cursor=True) as cur:
        cur.execute("SELECT * FROM matchoff_runs WHERE id = %s", (run_id,))
        row = cur.fetchone()

    if not row:
        raise ValueError(f"Matchoff run not found: {run_id}")
    return dict(row)


def export_and_store(
    module: str,
    records: list,
    columns: list,
    file_name: str,
    export_date: str,
    data_type: str = None,
) -> dict:
    container = _get_container_name(module)
    total = len(records)

    parts_raw = build_excel_parts(
        records=records,
        columns=columns,
        base_file_name=file_name,
        max_rows=EXCEL_MAX_ROWS,
    )

    parts_meta = []
    for buf, part_fname, row_count in parts_raw:
        blob_name = f"{export_date}/{part_fname}"
        upload_blob(container, blob_name, buf)

        row_id = save_export_metadata(
            module=module,
            file_name=part_fname,
            blob_name=blob_name,
            record_count=row_count,
            export_date=export_date,
            data_type=data_type,
        )

        parts_meta.append({
            "id": row_id,
            "file_name": part_fname,
            "record_count": row_count,
            "export_date": export_date,
            "module": module.upper(),
            "data_type": data_type,
        })
        print(f"[EXPORT] Uploaded {module} part {blob_name} ({row_count:,} rows)")

    return {
        "parts": parts_meta,
        "total_records": total,
        "file_count": len(parts_meta),
        "export_date": export_date,
        "module": module.upper(),
        "data_type": data_type,
        "record_count": total,
        "file_name": parts_meta[0]["file_name"] if parts_meta else file_name,
    }


def run_matchoff(
    source_container: str,
    date_folder: str,
    file_name: str,
    mob_col: str = "mob_num",
) -> dict:
    run_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    is_fri = source_container == AZURE_CONTAINER_FRI

    # 1. Read source Excel
    print(f"[MATCHOFF] Reading {source_container}/{date_folder}/{file_name}")
    input_rows, headers = read_excel_from_blob(source_container, date_folder, file_name)
    print(f"[MATCHOFF] {len(input_rows):,} rows, columns: {headers}")

    if mob_col not in headers:
        raise ValueError(
            f"Column '{mob_col}' not found in '{file_name}'. Available columns: {', '.join(headers)}"
        )

    if is_fri:
        if FRI_STATUS_COLUMN not in headers:
            eligible_rows = input_rows
            excluded_compliant_count = 0
        else:
            eligible_rows = [
                row for row in input_rows
                if str(row.get(FRI_STATUS_COLUMN, "")).strip() != FRI_EXCLUDE_STATUS
            ]
            excluded_compliant_count = len(input_rows) - len(eligible_rows)
            print(
                f"[MATCHOFF] FRI EXCLUSION: {excluded_compliant_count} '{FRI_EXCLUDE_STATUS}' "
                f"rows removed, {len(eligible_rows)} eligible rows remain"
            )
    else:
        eligible_rows = input_rows
        excluded_compliant_count = 0

    if not eligible_rows:
        raise ValueError(
            f"All {len(input_rows)} rows were excluded. No eligible records remain for match-off."
        )

    # 2. Extract + deduplicate mob numbers
    mob_numbers = [row.get(mob_col) for row in eligible_rows]
    unique_count = len({str(m).strip() for m in mob_numbers if m})
    print(f"[MATCHOFF] {unique_count:,} unique mob numbers to match")

    # 3. Chunked parallel match
    print(f"[MATCHOFF] Querying mobilecleanup: chunks={SQL_CHUNK_SIZE}, workers={SQL_MAX_WORKERS}")
    matched_map = match_mob_numbers(mob_numbers)
    print(f"[MATCHOFF] {len(matched_map):,} unique numbers matched")

    if not matched_map:
        raise ValueError(
            "Zero matches found. Verify mob_col name and that Phase 2 data has been loaded into mobilecleanup."
        )

    base_name = file_name.replace(".xlsx", "")
    matched_count_est = sum(
        1 for row in eligible_rows if str(row.get(mob_col, "")).strip() in matched_map
    )
    output_file_name = f"{base_name}_matched_{matched_count_est}_records.xlsx"
    print(f"[MATCHOFF] Building output Excel: {output_file_name}")

    buf, matched_count = build_matched_excel(
        input_rows=eligible_rows,
        headers=headers,
        mob_col=mob_col,
        matched_map=matched_map,
    )

    # 5. Upload
    output_blob_path = upload_matchoff_output(buf, date_folder, output_file_name)
    print(f"[MATCHOFF] Uploaded matchoff-output/{output_blob_path}")

    # 6. Metadata
    run_id = save_matchoff_metadata(
        source_container=source_container,
        source_date_folder=date_folder,
        source_file_name=file_name,
        output_blob_path=output_blob_path,
        input_row_count=len(input_rows),
        matched_count=matched_count,
        mob_col=mob_col,
        run_date=run_date,
    )

    # 7. 24-hour SAS URL
    sas_url = get_sas_download_url(
        container=AZURE_CONTAINER_MATCHOFF_OUTPUT,
        blob_name=output_blob_path,
        expiry_hours=24,
    )

    eligible_row_count = len(input_rows) - excluded_compliant_count

    print(
        f"[MATCHOFF] Done. input = {len(input_rows)}, "
        f"excluded_compliant={excluded_compliant_count}, "
        f"eligible={eligible_row_count}, "
        f"matched_count={matched_count:,}, run_id={run_id}"
    )

    return {
        "id": run_id,
        "source_container": source_container,
        "date_folder": date_folder,
        "source_file": file_name,
        "output_file": output_file_name,
        "input_row_count": len(input_rows),
        "excluded_compliant_count": excluded_compliant_count,
        "eligible_row_count": eligible_row_count,
        "matched_count": matched_count,
        "sas_url": sas_url,
        "run_date": run_date,
    }